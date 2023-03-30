import openpyxl
import os
from models import ProfessionStatistics


def save_statistics(data: list[ProfessionStatistics], filename:str = "result/статистика.xlsx") -> None:
    os.makedirs("result", exist_ok=True)
    book = openpyxl.Workbook()
    sheet = book.active
    writeTitles(sheet, data)
    column = 1
    for item in data:
        rowCount = 2
        for skill in item.Skills:
            if len(skill) < 2: continue
            sheet.cell(rowCount, column).value = skill.Name.lower()
            sheet.cell(rowCount, column+1).value = skill.Count
            rowCount += 1
        column += 2
    book.save(filename)
    


def writeTitles(sheet: openpyxl.worksheet.worksheet.Worksheet, data: list[ProfessionStatistics]):
    row = 1
    for item in data:
        sheet.cell(1, row).value = item.Profession        
        sheet.cell(1, row+1).value = "Повторений"
        row += 2